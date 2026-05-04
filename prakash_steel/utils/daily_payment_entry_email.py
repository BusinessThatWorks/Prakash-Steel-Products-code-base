from io import BytesIO

import frappe
from frappe.utils import date_diff, fmt_money, formatdate, now_datetime
from frappe.utils.data import escape_html
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


RECIPIENTS = [
	"pratikshya.gochhayat@clapgrow.com",
	"ritika@clapgrow.com",
	"beetashoke.chakraborty@clapgrow.com",
	"accounts@prakashsteel.com",
]


def send_daily_payment_entry_email():
	"""Send Payment Entry rows submitted today (by custom_submitted_time) within the configured window."""
	rows = _get_payment_entry_rows_for_today_till_11_am()

	if not rows:
		return

	report_date = formatdate(now_datetime().date(), "dd-MM-yyyy")
	table_rows, total_amount = _prepare_axis_bank_table_rows(rows)
	subject = f"Daily Payment Entry Report - {report_date}"
	body = _build_email_body(table_rows, total_amount, report_date)
	excel_bytes = _build_axis_bank_excel_bytes(table_rows, total_amount, report_date)
	attachment_fname = f"Daily_Payment_Entry_Report_{report_date.replace('-', '_')}.xlsx"

	frappe.sendmail(
		recipients=RECIPIENTS,
		subject=subject,
		message=body,
		attachments=[
			{
				"fname": attachment_fname,
				"fcontent": excel_bytes,
				"content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
			}
		],
	)


def _get_payment_entry_rows_for_today_till_11_am():
	"""Fetch Payment Entries submitted since yesterday 12:00 PM up to today 12:00 PM (scheduler window)."""
	return frappe.db.sql(
		"""
		SELECT
			pe.name AS payment_entry,
			pe.party_name,
			pe.posting_date AS received_date,
			pe.paid_amount,
			per.reference_name,
			si.posting_date AS sales_invoice_posting_date,
			c.customer_group
		FROM `tabPayment Entry` pe
		INNER JOIN `tabPayment Entry Reference` per
			ON per.parent = pe.name
		LEFT JOIN `tabCustomer` c
			ON c.name = pe.party AND IFNULL(pe.party_type, '') = 'Customer'
		LEFT JOIN `tabSales Invoice` si
			ON si.name = per.reference_name
		WHERE pe.docstatus = 1
		  AND pe.custom_submitted_time IS NOT NULL
		  AND pe.custom_submitted_time >= CONCAT(CURDATE(), ' 00:00:00')
		  AND pe.custom_submitted_time <  CONCAT(CURDATE(), ' 12:00:00')
		  AND IFNULL(per.reference_name, '') != ''
		  AND per.reference_doctype = 'Sales Invoice'
		ORDER BY pe.name, per.idx
		""",
		as_dict=True,
	)


def _prepare_axis_bank_table_rows(rows):
	"""Normalize SQL rows into display dicts for HTML and Excel (same columns as the email table)."""
	table_rows = []
	total_amount = 0

	for row in rows:
		invoice_posting_date = (
			formatdate(row.sales_invoice_posting_date, "dd-MM-yyyy") if row.sales_invoice_posting_date else ""
		)
		received_date = formatdate(row.received_date, "dd-MM-yyyy") if row.received_date else ""
		due_days = (
			date_diff(row.received_date, row.sales_invoice_posting_date)
			if row.sales_invoice_posting_date and row.received_date
			else ""
		)
		amount = row.paid_amount or 0
		total_amount += amount

		table_rows.append(
			{
				"particulars": row.party_name or "",
				"group": row.customer_group or "",
				"sales_invoice_no": row.reference_name or "",
				"invoice_date": invoice_posting_date,
				"received_date": received_date,
				"due_days": due_days,
				"amount": amount,
				"amount_display": fmt_money(amount, currency="INR"),
			}
		)

	return table_rows, total_amount


def _build_email_body(table_rows, total_amount, report_date_display):
	"""Build the HTML table body for payment entry rows."""
	table_rows_html = []

	for r in table_rows:
		due_cell = str(r["due_days"]) if r["due_days"] != "" else ""
		table_rows_html.append(
			f"""
			<tr>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;">{escape_html(r["particulars"])}</td>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;">{escape_html(r["group"])}</td>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;">{escape_html(r["sales_invoice_no"])}</td>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;">{escape_html(r["invoice_date"])}</td>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;">{escape_html(r["received_date"])}</td>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;text-align:center;">{due_cell}</td>
				<td style="padding:8px 10px;border:1px solid #cfcfcf;text-align:right;">{r["amount_display"]}</td>
			</tr>
			"""
		)

	return f"""
	<div style="font-family:Arial, Helvetica, sans-serif;color:#1f2937;">
		<p>Dear Team,</p>
		<p>Please find below the daily Payment Entry report for <b>{report_date_display}</b>. An Excel copy is attached.</p>

		<h3 style="margin:16px 0 8px 0;">AXIS BANK</h3>
		<table style="border-collapse:collapse;width:100%;font-size:13px;">
			<thead>
				<tr style="background-color:#f4f6f8;">
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:left;">Particulars</th>
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:left;">Group</th>
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:left;">Sales Invoice No.</th>
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:left;">Invoice Date</th>
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:left;">Received Date</th>
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:center;">Due Days</th>
					<th style="padding:9px 10px;border:1px solid #cfcfcf;text-align:right;">Amount</th>
				</tr>
			</thead>
			<tbody>
				{"".join(table_rows_html)}
			</tbody>
			<tfoot>
				<tr style="font-weight:700;background-color:#fafafa;">
					<td colspan="6" style="padding:9px 10px;border:1px solid #cfcfcf;text-align:right;">Total</td>
					<td style="padding:9px 10px;border:1px solid #cfcfcf;text-align:right;">{fmt_money(total_amount, currency="INR")}</td>
				</tr>
			</tfoot>
		</table>

		<p style="margin-top:16px;">Regards,<br>Prakash Steel ERP</p>
	</div>
	"""


def _build_axis_bank_excel_bytes(table_rows, total_amount, report_date_display):
	"""Build .xlsx bytes matching the AXIS BANK table in the email."""
	wb = Workbook()
	ws = wb.active
	ws.title = "AXIS BANK"

	header_font = Font(bold=True)
	header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
	ws.append(["Daily Payment Entry Report", report_date_display])
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
	ws["A1"].font = Font(bold=True, size=12)

	ws.append([])
	ws.append(["AXIS BANK"])
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)
	ws["A3"].font = Font(bold=True, size=11)

	headers = [
		"Particulars",
		"Group",
		"Sales Invoice No.",
		"Invoice Date",
		"Received Date",
		"Due Days",
		"Amount",
	]
	ws.append(headers)
	header_row = ws.max_row
	for col, _h in enumerate(headers, start=1):
		cell = ws.cell(row=header_row, column=col)
		cell.font = header_font
		cell.alignment = header_align

	for r in table_rows:
		ws.append(
			[
				r["particulars"],
				r["group"],
				r["sales_invoice_no"],
				r["invoice_date"],
				r["received_date"],
				r["due_days"] if r["due_days"] != "" else None,
				r["amount"],
			]
		)

	total_row = ws.max_row + 1
	ws.cell(row=total_row, column=1, value="Total")
	ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
	ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
	ws.cell(row=total_row, column=1).font = Font(bold=True)
	ws.cell(row=total_row, column=7, value=total_amount)
	ws.cell(row=total_row, column=7).font = Font(bold=True)
	ws.cell(row=total_row, column=7).number_format = "₹#,##0.00"

	for row in ws.iter_rows(min_row=header_row + 1, max_row=total_row - 1, min_col=7, max_col=7):
		for cell in row:
			cell.number_format = "₹#,##0.00"

	for col_letter in ("A", "C", "D", "E"):
		ws.column_dimensions[col_letter].width = 22
	ws.column_dimensions["B"].width = 18
	ws.column_dimensions["F"].width = 10
	ws.column_dimensions["G"].width = 16

	buf = BytesIO()
	wb.save(buf)
	return buf.getvalue()
