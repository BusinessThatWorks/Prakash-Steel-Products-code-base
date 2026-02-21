import json
from io import BytesIO
from typing import Any

import frappe
from frappe.utils import flt
from frappe.utils.file_manager import save_file
from frappe.utils.pdf import get_pdf


@frappe.whitelist()
def get_item_insight_data(
    from_date=None,
    to_date=None,
    item_code=None,
    item_grade=None,
    category_name=None,
    description_code=None,
    limit=None,
):
    """
    Fetch comprehensive item insight data using bulk queries for performance.
    Instead of querying per-item in a loop, all data is fetched in a handful
    of batch SQL queries and merged in Python.
    """

    # Build item filter
    item_filter: dict[str, object] = {"is_stock_item": 1}

    if item_code:
        item_filter["name"] = item_code
    if item_grade:
        item_filter["custom_grade"] = item_grade
    if category_name:
        item_filter["custom_category_name"] = category_name
    if description_code:
        item_filter["custom_desc_code"] = description_code

    # Get all matching items
    items = frappe.get_all(
        "Item",
        filters=item_filter,
        fields=[
            "name",
            "item_name",
            "item_code",
            "custom_grade as item_grade",
            "custom_category_name as category_name",
            "custom_desc_code",
        ],
        order_by="item_code",
        limit_page_length=int(limit) if limit else None,
    )

    if not items:
        return []

    # Classify items by type (Rolled vs Bright)
    all_codes = []
    rolled_codes = []
    bright_codes = []

    for item in items:
        code = item.name
        all_codes.append(code)
        desc = item.custom_desc_code or ""
        if "Bright Bar" in desc:
            bright_codes.append(code)
        else:
            rolled_codes.append(code)

    # ── Filter to items with transactions in date range ──
    if from_date and to_date:
        active_items = _get_items_with_transactions(
            all_codes, rolled_codes, bright_codes, from_date, to_date
        )
        # Keep only items that have at least one transaction
        items = [item for item in items if item.name in active_items]
        if not items:
            return []
        # Recompute code lists after filtering
        all_codes = []
        rolled_codes = []
        bright_codes = []
        for item in items:
            code = item.name
            all_codes.append(code)
            desc = item.custom_desc_code or ""
            if "Bright Bar" in desc:
                bright_codes.append(code)
            else:
                rolled_codes.append(code)

    # ── Bulk fetch all data (few queries instead of N × 6+) ──
    prod_date_map = _bulk_production_dates(
        rolled_codes, bright_codes, from_date, to_date
    )
    prod_qty_map = _bulk_production_qty(rolled_codes, bright_codes, from_date, to_date)
    sales_map, pending_so_map = _bulk_sales_data(all_codes, from_date, to_date)
    purchase_map, pending_po_map = _bulk_purchase_data(all_codes, from_date, to_date)
    inventory_map = _bulk_inventory_data(all_codes)

    # ── Merge results ──
    result = []
    for item in items:
        ic = item.name

        inv = inventory_map.get(ic, [])
        total_stock = sum(flt(w.get("stock_qty", 0), 2) for w in inv)
        total_committed = sum(flt(w.get("committed_stock", 0), 2) for w in inv)

        sd = sales_map.get(ic, {})
        pd = purchase_map.get(ic, {})

        item_data = {
            "item_code": item.item_code,
            "item_name": item.item_name or item.item_code,
            # Production
            "last_production_date": prod_date_map.get(ic),
            "last_production_quantity": prod_qty_map.get(ic, 0),
            # Sales
            "last_sales_party": sd.get("customer_name"),
            "last_sales_date": sd.get("posting_date"),
            "last_sales_quantity": flt(sd.get("qty", 0), 2),
            "last_sales_rate": flt(sd.get("rate", 0), 2),
            "pending_sales_order_qty": pending_so_map.get(ic, 0),
            # Purchase
            "last_purchase_party": pd.get("supplier_name"),
            "last_purchase_date": pd.get("posting_date"),
            "last_purchase_quantity": flt(pd.get("qty", 0), 2),
            "last_purchase_rate": flt(pd.get("rate", 0), 2),
            "pending_purchase_order_qty": pending_po_map.get(ic, 0),
            # Inventory
            "warehouse_stock": inv,
            "total_stock_on_hand": flt(total_stock, 2),
            "committed_stock": flt(total_committed, 2),
            "projected_qty": flt(total_stock - total_committed, 2),
        }

        if getattr(item, "item_grade", None):
            item_data["item_grade"] = item.item_grade
        if getattr(item, "category_name", None):
            item_data["category_name"] = item.category_name

        result.append(item_data)

    return result


# ───────────────────────────────────────────────────────────────────────────
# Bulk helper functions
# ───────────────────────────────────────────────────────────────────────────


def _get_items_with_transactions(
    all_codes, rolled_codes, bright_codes, from_date, to_date
):
    """
    Find item codes that have at least one transaction (Sales, Purchase, or
    Production) within the given date range.  Uses a single UNION ALL query
    for speed.  Returns a set of item_codes.
    """
    if not all_codes or not from_date or not to_date:
        return set(all_codes)

    codes_tuple = tuple(all_codes)
    parts = []
    params: list = []

    # Sales Invoice
    parts.append(
        "SELECT DISTINCT sii.item_code "
        "FROM `tabSales Invoice Item` sii "
        "INNER JOIN `tabSales Invoice` si ON si.name = sii.parent "
        "WHERE sii.item_code IN %s AND si.docstatus = 1 "
        "AND si.posting_date BETWEEN %s AND %s"
    )
    params.extend([codes_tuple, from_date, to_date])

    # Purchase Invoice
    parts.append(
        "SELECT DISTINCT pii.item_code "
        "FROM `tabPurchase Invoice Item` pii "
        "INNER JOIN `tabPurchase Invoice` pi ON pi.name = pii.parent "
        "WHERE pii.item_code IN %s AND pi.docstatus = 1 "
        "AND pi.posting_date BETWEEN %s AND %s"
    )
    params.extend([codes_tuple, from_date, to_date])

    # Finish Weight (rolled bars)
    if rolled_codes:
        parts.append(
            "SELECT DISTINCT item_code "
            "FROM `tabFinish Weight` "
            "WHERE item_code IN %s AND docstatus = 1 "
            "AND posting_date BETWEEN %s AND %s"
        )
        params.extend([tuple(rolled_codes), from_date, to_date])

    # Bright Bar Production (bright bars)
    if bright_codes:
        parts.append(
            "SELECT DISTINCT finished_good AS item_code "
            "FROM `tabBright Bar Production` "
            "WHERE finished_good IN %s AND docstatus = 1 "
            "AND production_date BETWEEN %s AND %s"
        )
        params.extend([tuple(bright_codes), from_date, to_date])

    query = " UNION ALL ".join(parts)
    rows = frappe.db.sql(query, tuple(params), as_dict=True)
    return {r.item_code for r in rows}


def _bulk_production_dates(rolled_codes, bright_codes, from_date, to_date):
    """
    Get the latest production date per item.
    - Rolled bars → Finish Weight.posting_date
    - Bright bars → Bright Bar Production.production_date
    Returns {item_code: date}.
    """
    result = {}

    # Rolled bars: Finish Weight
    if rolled_codes:
        query = """
            SELECT item_code, posting_date
            FROM `tabFinish Weight`
            WHERE item_code IN %s
                AND docstatus = 1
        """
        params: list = [tuple(rolled_codes)]
        if from_date and to_date:
            query += " AND posting_date BETWEEN %s AND %s"
            params.extend([from_date, to_date])
        query += " ORDER BY posting_date DESC, creation DESC"

        rows = frappe.db.sql(query, tuple(params), as_dict=True)
        for r in rows:
            if r.item_code not in result:
                result[r.item_code] = r.posting_date

    # Bright bars: Bright Bar Production
    if bright_codes:
        query = """
            SELECT finished_good AS item_code, production_date
            FROM `tabBright Bar Production`
            WHERE finished_good IN %s
                AND docstatus = 1
        """
        params = [tuple(bright_codes)]
        if from_date and to_date:
            query += " AND production_date BETWEEN %s AND %s"
            params.extend([from_date, to_date])
        query += " ORDER BY production_date DESC, creation DESC"

        rows = frappe.db.sql(query, tuple(params), as_dict=True)
        for r in rows:
            if r.item_code not in result:
                result[r.item_code] = r.production_date

    return result


def _bulk_production_qty(rolled_codes, bright_codes, from_date, to_date):
    """
    Get the latest production quantity per item.
    - Rolled bars → Finish Weight.finish_weight
    - Bright bars → Bright Bar Production.fg_weight
    Returns {item_code: qty}.
    """
    result = {}

    # Rolled bars: Finish Weight
    if rolled_codes:
        query = """
            SELECT item_code, finish_weight
            FROM `tabFinish Weight`
            WHERE item_code IN %s
                AND docstatus = 1
        """
        params: list = [tuple(rolled_codes)]
        if from_date and to_date:
            query += " AND posting_date BETWEEN %s AND %s"
            params.extend([from_date, to_date])
        query += " ORDER BY posting_date DESC, creation DESC"

        rows = frappe.db.sql(query, tuple(params), as_dict=True)
        for r in rows:
            if r.item_code not in result:
                result[r.item_code] = flt(r.finish_weight, 2)

    # Bright bars: Bright Bar Production
    if bright_codes:
        query = """
            SELECT finished_good AS item_code, fg_weight
            FROM `tabBright Bar Production`
            WHERE finished_good IN %s
                AND docstatus = 1
        """
        params = [tuple(bright_codes)]
        if from_date and to_date:
            query += " AND production_date BETWEEN %s AND %s"
            params.extend([from_date, to_date])
        query += " ORDER BY production_date DESC, creation DESC"

        rows = frappe.db.sql(query, tuple(params), as_dict=True)
        for r in rows:
            if r.item_code not in result:
                result[r.item_code] = flt(r.fg_weight, 2)

    return result


def _bulk_sales_data(item_codes, from_date, to_date):
    """
    Bulk-fetch latest Sales Invoice data and pending SO qty.
    Returns (sales_map, pending_so_map).
    """
    sales_map: dict = {}
    pending_so_map: dict = {}

    if not item_codes:
        return sales_map, pending_so_map

    codes_tuple = tuple(item_codes)

    # Latest Sales Invoice per item
    query = """
        SELECT sii.item_code, si.customer_name, sii.qty, sii.rate, si.posting_date
        FROM `tabSales Invoice Item` sii
        INNER JOIN `tabSales Invoice` si ON si.name = sii.parent
        WHERE sii.item_code IN %s
            AND si.docstatus = 1
    """
    params: list = [codes_tuple]
    if from_date and to_date:
        query += " AND si.posting_date BETWEEN %s AND %s"
        params.extend([from_date, to_date])
    query += " ORDER BY si.posting_date DESC, si.posting_time DESC, si.creation DESC"

    rows = frappe.db.sql(query, tuple(params), as_dict=True)
    for r in rows:
        if r.item_code not in sales_map:
            sales_map[r.item_code] = {
                "customer_name": r.customer_name,
                "qty": r.qty,
                "rate": r.rate,
                "posting_date": r.posting_date,
            }

    # Pending SO qty (aggregated per item)
    pending_rows = frappe.db.sql(
        """
        SELECT soi.item_code, SUM(soi.qty - IFNULL(soi.delivered_qty, 0)) AS pending_qty
        FROM `tabSales Order Item` soi
        INNER JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE soi.item_code IN %s
            AND so.status NOT IN ('Stopped', 'On Hold', 'Closed', 'Cancelled', 'Completed')
            AND so.docstatus = 1
            AND (soi.qty - IFNULL(soi.delivered_qty, 0)) > 0
        GROUP BY soi.item_code
    """,
        (codes_tuple,),
        as_dict=True,
    )
    for r in pending_rows:
        pending_so_map[r.item_code] = flt(r.pending_qty, 2)

    return sales_map, pending_so_map


def _bulk_purchase_data(item_codes, from_date, to_date):
    """
    Bulk-fetch latest Purchase Invoice data and pending PO qty.
    Returns (purchase_map, pending_po_map).
    """
    purchase_map: dict = {}
    pending_po_map: dict = {}

    if not item_codes:
        return purchase_map, pending_po_map

    codes_tuple = tuple(item_codes)

    # Latest Purchase Invoice per item
    query = """
        SELECT pii.item_code, pi.supplier_name, pii.qty, pii.rate, pi.posting_date
        FROM `tabPurchase Invoice Item` pii
        INNER JOIN `tabPurchase Invoice` pi ON pi.name = pii.parent
        WHERE pii.item_code IN %s
            AND pi.docstatus = 1
    """
    params: list = [codes_tuple]
    if from_date and to_date:
        query += " AND pi.posting_date BETWEEN %s AND %s"
        params.extend([from_date, to_date])
    query += " ORDER BY pi.posting_date DESC, pi.posting_time DESC, pi.creation DESC"

    rows = frappe.db.sql(query, tuple(params), as_dict=True)
    for r in rows:
        if r.item_code not in purchase_map:
            purchase_map[r.item_code] = {
                "supplier_name": r.supplier_name,
                "qty": r.qty,
                "rate": r.rate,
                "posting_date": r.posting_date,
            }

    # Pending PO qty (aggregated per item)
    pending_rows = frappe.db.sql(
        """
        SELECT poi.item_code, SUM(poi.qty - IFNULL(poi.received_qty, 0)) AS pending_qty
        FROM `tabPurchase Order Item` poi
        INNER JOIN `tabPurchase Order` po ON po.name = poi.parent
        WHERE poi.item_code IN %s
            AND po.status NOT IN ('Stopped', 'On Hold', 'Closed', 'Cancelled', 'Completed')
            AND po.docstatus = 1
            AND (poi.qty - IFNULL(poi.received_qty, 0)) > 0
        GROUP BY poi.item_code
    """,
        (codes_tuple,),
        as_dict=True,
    )
    for r in pending_rows:
        pending_po_map[r.item_code] = flt(r.pending_qty, 2)

    return purchase_map, pending_po_map


def _bulk_inventory_data(item_codes):
    """
    Bulk-fetch warehouse-wise stock for all items.
    Returns {item_code: [{warehouse, stock_qty, committed_stock, projected_qty}]}.
    """
    if not item_codes:
        return {}

    excluded_warehouses = ["Rejected Warehouse"]

    rows = frappe.db.sql(
        """
        SELECT item_code, warehouse,
               SUM(actual_qty) AS stock_qty,
               SUM(IFNULL(reserved_qty, 0)) AS committed_stock
        FROM `tabBin`
        WHERE item_code IN %s
            AND warehouse NOT IN %s
            AND (actual_qty != 0 OR IFNULL(reserved_qty, 0) != 0)
        GROUP BY item_code, warehouse
        ORDER BY item_code, warehouse
    """,
        (tuple(item_codes), tuple(excluded_warehouses)),
        as_dict=True,
    )

    result: dict = {}
    for r in rows:
        stock_qty = flt(r.stock_qty, 2)
        committed_stock = flt(r.committed_stock, 2)
        projected_qty = flt(stock_qty - committed_stock, 2)

        if r.item_code not in result:
            result[r.item_code] = []

        result[r.item_code].append(
            {
                "warehouse": r.warehouse,
                "stock_qty": stock_qty,
                "committed_stock": committed_stock,
                "projected_qty": projected_qty,
            }
        )

    return result


# ───────────────────────────────────────────────────────────────────────────
# Other API endpoints (unchanged)
# ───────────────────────────────────────────────────────────────────────────


@frappe.whitelist()
def search_items(query, limit=20):
    """
    Search for items by item code or item name
    Returns items with their name, item_name, and item_group
    """
    limit = int(limit)

    # If no query, return top items
    if not query or query.strip() == "":
        items = frappe.db.sql(
            """
			SELECT 
				name,
				item_name,
				item_group
			FROM `tabItem`
			WHERE is_stock_item = 1
			ORDER BY name
			LIMIT %(limit)s
		""",
            {"limit": limit},
            as_dict=True,
        )
    else:
        # Search by item code or item name
        items = frappe.db.sql(
            """
			SELECT 
				name,
				item_name,
				item_group
			FROM `tabItem`
			WHERE is_stock_item = 1
				AND (
					name LIKE %(query)s
					OR item_name LIKE %(query)s
				)
			ORDER BY 
				CASE 
					WHEN name LIKE %(exact_query)s THEN 1
					WHEN item_name LIKE %(exact_query)s THEN 2
					ELSE 3
				END,
				name
			LIMIT %(limit)s
		""",
            {"query": f"%{query}%", "exact_query": f"{query}%", "limit": limit},
            as_dict=True,
        )

    return items


@frappe.whitelist()
def search_item_categories(
    doctype: str,
    txt: str,
    searchfield: str,
    start: int,
    page_len: int,
    filters: dict[str, Any] | None = None,
):
    """
    Link field query: return Item Categories restricted to Items having the given grade.

    This is used to populate the Category Name suggestions based on the selected Item Grade.
    """
    filters = filters or {}
    item_grade = filters.get("item_grade")

    if not item_grade:
        # If no grade selected, fall back to standard Item Category search
        return frappe.db.sql(
            """
			SELECT 
				name,
				category_name
			FROM `tabItem Category`
			WHERE
				(name LIKE %(txt)s OR category_name LIKE %(txt)s)
			ORDER BY name
			LIMIT %(page_len)s OFFSET %(start)s
		""",
            {
                "txt": f"%{txt}%",
                "page_len": page_len,
                "start": start,
            },
        )

    return frappe.db.sql(
        """
		SELECT DISTINCT
			ic.name,
			ic.category_name
		FROM `tabItem Category` ic
		INNER JOIN `tabItem` i ON i.custom_category_name = ic.name
		WHERE
			i.custom_grade = %(item_grade)s
			AND (
				ic.name LIKE %(txt)s
				OR ic.category_name LIKE %(txt)s
			)
		ORDER BY ic.name
		LIMIT %(page_len)s OFFSET %(start)s
	""",
        {
            "item_grade": item_grade,
            "txt": f"%{txt}%",
            "page_len": page_len,
            "start": start,
        },
    )


@frappe.whitelist()
def export_item_insight_excel(filters: str | None = None) -> dict[str, Any]:
    """Export item insight data to Excel and return file URL."""

    filters_dict: dict[str, Any] = {}
    if filters:
        try:
            filters_dict = json.loads(filters)
        except Exception:
            frappe.throw("Invalid filters JSON")

    data = get_item_insight_data(**filters_dict)

    if not data:
        frappe.throw("No data to export")

    # Define column headers
    column_labels = [
        "Item Code",
        "Item Name",
        "Item Grade",
        "Category Name",
        "Last Production Date",
        "Last Production Qty",
        "Last Sales Party",
        "Last Sales Date",
        "Last Sales Qty",
        "Last Sales Rate",
        "Pending SO Qty",
        "Last Purchase Party",
        "Last Purchase Date",
        "Last Purchase Qty",
        "Last Purchase Rate",
        "Pending PO Qty",
        "Warehouse",
        "Stock Qty",
        "Committed Stock",
        "Projected Qty",
        "Total Stock On Hand",
    ]

    # Flatten data to include warehouse rows
    rows = []
    for item in data:
        warehouse_stock = item.get("warehouse_stock", [])

        if warehouse_stock:
            # Create a row for each warehouse
            for wh in warehouse_stock:
                row = [
                    item.get("item_code", ""),
                    item.get("item_name", ""),
                    item.get("item_grade", ""),
                    item.get("category_name", ""),
                    (
                        frappe.format(
                            item.get("last_production_date"), {"fieldtype": "Date"}
                        )
                        if item.get("last_production_date")
                        else ""
                    ),
                    flt(item.get("last_production_quantity", 0), 2),
                    item.get("last_sales_party", ""),
                    (
                        frappe.format(
                            item.get("last_sales_date"), {"fieldtype": "Date"}
                        )
                        if item.get("last_sales_date")
                        else ""
                    ),
                    flt(item.get("last_sales_quantity", 0), 2),
                    flt(item.get("last_sales_rate", 0), 2),
                    flt(item.get("pending_sales_order_qty", 0), 2),
                    item.get("last_purchase_party", ""),
                    (
                        frappe.format(
                            item.get("last_purchase_date"), {"fieldtype": "Date"}
                        )
                        if item.get("last_purchase_date")
                        else ""
                    ),
                    flt(item.get("last_purchase_quantity", 0), 2),
                    flt(item.get("last_purchase_rate", 0), 2),
                    flt(item.get("pending_purchase_order_qty", 0), 2),
                    wh.get("warehouse", ""),
                    flt(wh.get("stock_qty", 0), 2),
                    flt(wh.get("committed_stock", 0), 2),
                    flt(wh.get("projected_qty", 0), 2),
                    flt(item.get("total_stock_on_hand", 0), 2),
                ]
                rows.append(row)
        else:
            # No warehouses, create single row with empty warehouse fields
            row = [
                item.get("item_code", ""),
                item.get("item_name", ""),
                item.get("item_grade", ""),
                item.get("category_name", ""),
                (
                    frappe.format(
                        item.get("last_production_date"), {"fieldtype": "Date"}
                    )
                    if item.get("last_production_date")
                    else ""
                ),
                flt(item.get("last_production_quantity", 0), 2),
                item.get("last_sales_party", ""),
                (
                    frappe.format(item.get("last_sales_date"), {"fieldtype": "Date"})
                    if item.get("last_sales_date")
                    else ""
                ),
                flt(item.get("last_sales_quantity", 0), 2),
                flt(item.get("last_sales_rate", 0), 2),
                flt(item.get("pending_sales_order_qty", 0), 2),
                item.get("last_purchase_party", ""),
                (
                    frappe.format(item.get("last_purchase_date"), {"fieldtype": "Date"})
                    if item.get("last_purchase_date")
                    else ""
                ),
                flt(item.get("last_purchase_quantity", 0), 2),
                flt(item.get("last_purchase_rate", 0), 2),
                flt(item.get("pending_purchase_order_qty", 0), 2),
                "",  # warehouse
                "",  # stock_qty
                "",  # committed_stock
                "",  # projected_qty
                flt(item.get("total_stock_on_hand", 0), 2),
            ]
            rows.append(row)

    # Build the xlsx file with header row + data rows using openpyxl (built-in Frappe dep)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Item Insight"

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="D7E4BC", end_color="D7E4BC", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_num, header in enumerate(column_labels, start=1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Write data rows
    for row_num, row_data in enumerate(rows, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    xlsx_buffer = BytesIO()
    workbook.save(xlsx_buffer)
    xlsx_buffer.seek(0)

    file_name = "Item Insight Dashboard.xlsx"
    saved_file = save_file(
        fname=file_name,
        content=xlsx_buffer.getvalue(),
        dt=None,
        dn=None,
        is_private=1,
    )

    return {"file_url": saved_file.file_url}


@frappe.whitelist()
def export_item_insight_pdf(filters: str | None = None) -> dict[str, Any]:
    """Export item insight data to PDF and return file URL."""

    filters_dict: dict[str, Any] = {}
    if filters:
        try:
            filters_dict = json.loads(filters)
        except Exception:
            frappe.throw("Invalid filters JSON")

    data = get_item_insight_data(**filters_dict)

    if not data:
        frappe.throw("No data to export")

    # Build simple HTML table for PDF
    columns = [
        ("item_code", "Item Code"),
        ("item_name", "Item Name"),
        ("item_grade", "Item Grade"),
        ("category_name", "Category Name"),
        ("last_production_date", "Last Production Date"),
        ("last_production_quantity", "Last Production Qty"),
        ("last_sales_party", "Last Sales Party"),
        ("last_sales_date", "Last Sales Date"),
        ("last_sales_quantity", "Last Sales Qty"),
        ("last_sales_rate", "Last Sales Rate"),
        ("pending_sales_order_qty", "Pending SO Qty"),
        ("last_purchase_party", "Last Purchase Party"),
        ("last_purchase_date", "Last Purchase Date"),
        ("last_purchase_quantity", "Last Purchase Qty"),
        ("last_purchase_rate", "Last Purchase Rate"),
        ("pending_purchase_order_qty", "Pending PO Qty"),
        ("committed_stock", "Committed Stock"),
        ("projected_qty", "Projected Qty"),
        ("total_stock_on_hand", "Total Stock On Hand"),
    ]

    header_html = "".join(f"<th>{frappe._(label)}</th>" for _field, label in columns)

    body_rows = []
    for row in data:
        cells = []
        for field, _label in columns:
            value = row.get(field)
            cells.append(f"<td>{frappe.format(value)}</td>")
        body_rows.append(f"<tr>{''.join(cells)}</tr>")

    html = f"""
        <h3 style="text-align:center;">Item Insight Dashboard</h3>
        <table class="table table-bordered" style="width:100%;border-collapse:collapse;font-size:9pt;">
            <thead>
                <tr>{header_html}</tr>
            </thead>
            <tbody>
                {''.join(body_rows)}
            </tbody>
        </table>
    """

    pdf_content = get_pdf(html)

    file_name = "Item Insight Dashboard.pdf"
    saved_file = save_file(
        fname=file_name,
        content=pdf_content,
        dt=None,
        dn=None,
        is_private=1,
    )

    return {"file_url": saved_file.file_url}
